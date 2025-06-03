import os
import difflib
import xlsxwriter
import re
import openpyxl
import xlrd

def highlight_diff_in_line(line1, line2):
    """Return a flattened list of strings and formats for differing characters."""
    d = difflib.Differ()
    diff = list(d.compare(line1, line2))
    formatted_line = []

    for char in diff:
        if char.startswith("  "):  # No difference
            formatted_line.append(("black", char[2:]))  # Normal text in black
        elif char.startswith("- "):  # Difference in File1
            formatted_line.append(("red", char[2:]))  # Differing text in red
    return formatted_line

def write_rich_string(worksheet, row, col, formatted_line, workbook):
    """Write a formatted line into a cell with rich text formatting."""
    rich_text = []
    for color, text in formatted_line:
        fmt = workbook.add_format({"color": color})
        rich_text.append(fmt)
        rich_text.append(text)
    worksheet.write_rich_string(row, col, *rich_text)

def search_in_file(target_file, search_phrase, start_line=0):
    """Search for a phrase in the target file line by line starting from a specific line."""
    try:
        if not os.path.isfile(target_file):
            return f"Error: {target_file} not found", None, None

        search_phrase = search_phrase.strip().lower()
        with open(target_file, 'r', encoding='utf-8') as file:
            line_number = 1
            matched_lines = []
            for line in file:
                if line_number >= start_line:
                    normalized_line = ' '.join(line.strip().split()).lower()
                    if search_phrase in normalized_line:
                        matched_lines.append(line)
                        for _ in range(4):
                            next_line = next(file, None)
                            if next_line:
                                matched_lines.append(next_line)
                        return "Yes", matched_lines, line_number
                line_number += 1
        return "No", None, None
    except Exception as e:
        return f"Error: {e}", None, None

def contains_bit_word(line):
    """Check if the line contains the word 'bit' (case insensitive)."""
    return 'bit' in line.lower()

def filter_final_string(mdut_data):
    """Filter MDUT data and return valid strings for the 'Final String' column, including strings within double quotes."""
    valid_string_pattern = r"^[a-zA-Z0-9_]+$"
    invalid_words = {"fld", "linked", "from"}
    valid_strings = []

    if "from Register:" in mdut_data:
        parts = mdut_data.split("from Register:")[1:]
        for part in parts:
            part = part.strip()
            words = part.split()
            for word in words:
                if re.match(valid_string_pattern, word) and not word.isdigit() and "bit" not in word.lower() and word.lower() not in invalid_words:
                    valid_strings.append(word)
                    break
    else:
        words = mdut_data.strip().split()
        for word in words:
            if re.match(valid_string_pattern, word) and not word.isdigit() and "bit" not in word.lower() and word.lower() not in invalid_words:
                valid_strings.append(word)

    # Extract strings enclosed in double quotes
    double_quote_strings = re.findall(r'"(.*?)"', mdut_data)
    valid_strings.extend(double_quote_strings)

    return valid_strings

def find_testname_column_and_rows(excel_file_path, final_strings):
    """Search for strings in the 'Flow' tab and find 'TestName' column."""
    try:
        row_matches = {}
        testname_col = None
        sheet_data = None

        if excel_file_path.endswith('.xls'):
            workbook = xlrd.open_workbook(excel_file_path)
            sheet = None
            for sheet_name in workbook.sheet_names():
                if sheet_name.lower() == "flow":
                    sheet = workbook.sheet_by_name(sheet_name)
                    break

            if not sheet:
                raise ValueError("Flow tab not found in the Excel file.")

            for col_idx in range(sheet.ncols):
                for row_idx in range(min(3, sheet.nrows)):
                    if str(sheet.cell_value(row_idx, col_idx)).strip().lower() in {
                        "testname", "test name"
                    }:
                        testname_col = col_idx
                        break
                if testname_col is not None:
                    break

            if testname_col is None:
                raise ValueError("TestName column not found in the first 3 rows.")

            sheet_data = {"type": "xls", "sheet": sheet}

            for string in final_strings:
                for row_idx in range(sheet.nrows):
                    row_values = sheet.row_values(row_idx)
                    if any(string in str(cell) for cell in row_values):
                        row_matches[string] = row_idx + 1

        elif excel_file_path.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = None
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() == "flow":
                    sheet = workbook[sheet_name]
                    break

            if not sheet:
                raise ValueError("Flow tab not found in the Excel file.")

            for row in sheet.iter_rows(min_row=1, max_row=3):
                for cell in row:
                    if cell.value and str(cell.value).strip().lower() in {
                        "testname", "test name"
                    }:
                        testname_col = cell.column
                        break
                if testname_col:
                    break

            if not testname_col:
                raise ValueError("TestName column not found in the first 3 rows.")

            sheet_data = {"type": "xlsx", "sheet": sheet}

            for string in final_strings:
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if any(string in str(cell) for cell in row if cell):
                        row_matches[string] = row_idx

        else:
            raise ValueError("Unsupported file format. Only .xls and .xlsx are supported.")

        return row_matches, testname_col, sheet_data
    except Exception as e:
        print(f"Error while processing the Excel file: {e}")
        return {}, None, None

def compare_and_generate_report(file1_path, file2_path, target_file, excel_file_path, worksheet, workbook):
    """Comparison and result writing logic."""
    try:
        with open(file1_path, 'r', encoding='utf-8') as file1, open(file2_path, 'r', encoding='utf-8') as file2:
            file1_lines = file1.readlines()
            file2_lines = file2.readlines()

        worksheet.write_row(
            0, 0,
            ["Line #", "File1 Line", "File2 Line", "MDUT Data", "Final String", "TRD Data"]
        )

        line_number = 1
        search_start_line = 0

        for i, (line1, line2) in enumerate(zip(file1_lines, file2_lines), start=1):
            if line1.strip() != line2.strip():
                file1_diff = highlight_diff_in_line(line1.strip(), line2.strip())
                file2_diff = highlight_diff_in_line(line2.strip(), line1.strip())

                worksheet.write(line_number, 0, i)
                write_rich_string(worksheet, line_number, 1, file1_diff, workbook)
                write_rich_string(worksheet, line_number, 2, file2_diff, workbook)

                first_word = next((word for word in line1.split() if len(word) > 1), None)
                search_phrase_1 = f"|> Register {first_word}" if first_word else "|> Register N/A"
                search_phrase_2 = f"[DBG]: reg \"{first_word}\"" if first_word else "[DBG]: reg N/A"

                # First Query: Searching for "|> Register <extracted word>"
                found_query_1, matched_lines_query_1, found_line_query_1 = search_in_file(target_file, search_phrase_1, search_start_line)
                context_query_1 = []

                if matched_lines_query_1:
                    for line in matched_lines_query_1[1:]:
                        if contains_bit_word(line):
                            context_query_1.append(line.strip())
                        else:
                            break

                    if context_query_1:
                        worksheet.write(line_number, 3, "\n".join(context_query_1))

                # Second Query: If the first query yields no result, search for "[DBG]: reg '<extracted word>'"
                if not context_query_1:
                    found_query_2, matched_lines_query_2, found_line_query_2 = search_in_file(target_file, search_phrase_2, search_start_line)
                    context_query_2 = []

                    if matched_lines_query_2:
                        for line in matched_lines_query_2[1:]:
                            if "fld" in line.lower():
                                context_query_2.append(line.strip())
                            else:
                                break

                        if context_query_2:
                            worksheet.write(line_number, 3, "\n".join(context_query_2))

                # Use whichever context has been found (priority: first query > second query)
                final_strings = filter_final_string("\n".join(context_query_1 if context_query_1 else context_query_2))
                if final_strings:
                    worksheet.write(line_number, 4, "\n".join(final_strings))

                if final_strings and excel_file_path:
                    row_matches, testname_col, sheet_data = find_testname_column_and_rows(excel_file_path, final_strings)
                    trd_data = []
                    for string in final_strings:
                        if string in row_matches:
                            row = row_matches[string]
                            if sheet_data["type"] == "xlsx":
                                testname_value = sheet_data["sheet"].cell(row=row, column=testname_col).value
                            else:
                                testname_value = sheet_data["sheet"].cell_value(row - 1, testname_col)

                            if testname_value:
                                trd_data.append(f"TestName - {testname_value} found for Final String: '{string}'")
                        else:
                            trd_data.append("No MDUT label found in TRD")

                    if trd_data:
                        worksheet.write(line_number, 5, "\n".join(trd_data))

                line_number += 1

    except FileNotFoundError as e:
        print(f"Error: {e}")
    except UnicodeDecodeError as e:
        print(f"Unicode error: {e}")
        print("Try opening the file with a different encoding (e.g., 'latin-1').")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    print("Do you want to process both Sort-1 and Sort-2, or only one flow? (Both / Sort-1 / Sort-2)")
    choice = input("Enter your choice: ").strip().lower()

    target_file = input("Enter the path to the MDUT log file(File in which SMROWs are mapped to MDUT labels. Can be loading logs or read smr to mdut test logs--format- txt: ").strip()
    excel_file_path = input("Enter the path to the Excel File for TRD Data: ").strip()

    workbook = xlsxwriter.Workbook("comparison_report.xlsx")

    if choice == "both":
        sort1_file1 = input("Input Imprint verify content before Sort-1 flow run--file1--format- txt: ").strip()
        sort1_file2 = input("Input Imprint verify content after Sort-1 flow run--file2--format- txt: ").strip()
        sort2_file1 = input("Input Imprint verify content before Sort-2 flow run--file1--format- txt: ").strip()
        sort2_file2 = input("Input Imprint verify content after Sort-2 flow run--file2--format- txt: ").strip()

        worksheet1 = workbook.add_worksheet("Sort-1")
        compare_and_generate_report(sort1_file1, sort1_file2, target_file, excel_file_path, worksheet1, workbook)

        worksheet2 = workbook.add_worksheet("Sort-2")
        compare_and_generate_report(sort2_file1, sort2_file2, target_file, excel_file_path, worksheet2, workbook)

    elif choice in {"sort-1", "sort-2"}:
        file1 = input(f"Input Imprint verify content before {choice.capitalize()} flow run--file1--format- txt: ").strip()
        file2 = input(f"Input Imprint verify content after {choice.capitalize()} flow run--file2--format- txt: ").strip()

        worksheet = workbook.add_worksheet(choice.capitalize())
        compare_and_generate_report(file1, file2, target_file, excel_file_path, worksheet, workbook)

    else:
        print("Invalid choice. Exiting.")

    workbook.close()
    print("Comparison report has been successfully generated.")
